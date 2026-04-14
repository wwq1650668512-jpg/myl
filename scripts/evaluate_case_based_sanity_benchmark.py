from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BENCHMARK_XLSX = ROOT / "case_based_sanity_benchmark.xlsx"
FUSION_DIR = ROOT / "predictions" / "evaluation" / "fusion_comparison"
PROFILE_EVIDENCE_PATH = FUSION_DIR / "profile_evidence.json"
RESULTS_CSV_PATH = FUSION_DIR / "case_based_results.csv"

RAW_MODE = "raw_only"
MECH_MODE = "mechanism_only"
FINAL_MODE = "weighted_0.3_0.7"

MODEL_VERSION = "mechanism_layer_v1 + fusion_compare(raw/mechanism/0.3_0.7) + drug_profile_gating_v1"

DRUG_TO_FILE = {
    "Rifaximin": "rifaximin.csv",
    "Vancomycin": "vancomycin.csv",
    "Lubiprostone": "lubiprostone.csv",
    "Metronidazole": "metronidazole.csv",
}

RESULT_COLUMNS = [
    "assertion_id",
    "Drug",
    "Model version",
    "Run date",
    "Raw evidence",
    "Mechanism evidence",
    "Final evidence",
    "Status",
    "Confidence",
    "Key supporting output",
    "Key conflicting output",
    "Reviewer notes",
]


@dataclass
class AssertionResult:
    assertion_id: str
    drug: str
    raw_evidence: str
    mechanism_evidence: str
    final_evidence: str
    status: str
    confidence: str
    supporting: str
    conflicting: str
    notes: str

    def to_row(self) -> dict[str, Any]:
        return {
            "assertion_id": self.assertion_id,
            "Drug": self.drug,
            "Model version": MODEL_VERSION,
            "Run date": datetime.now().strftime("%Y-%m-%d"),
            "Raw evidence": self.raw_evidence,
            "Mechanism evidence": self.mechanism_evidence,
            "Final evidence": self.final_evidence,
            "Status": self.status,
            "Confidence": self.confidence,
            "Key supporting output": self.supporting,
            "Key conflicting output": self.conflicting,
            "Reviewer notes": self.notes,
        }


def _load_drug_tables() -> dict[str, pd.DataFrame]:
    tables: dict[str, pd.DataFrame] = {}
    for drug, file_name in DRUG_TO_FILE.items():
        path = FUSION_DIR / file_name
        if not path.exists():
            raise FileNotFoundError(f"Missing fusion result file: {path}")
        table = pd.read_csv(path)
        required = {
            "fusion_mode",
            "disease_name",
            "final_score",
            "rank",
            "pro_inflammatory_score",
            "toxin_risk_score",
            "pathobiont_load",
            "barrier_protection_score",
            "butyrate_support_score",
        }
        missing = sorted(required - set(table.columns))
        if missing:
            raise ValueError(f"{path} missing required columns: {missing}")
        tables[drug] = table
    return tables


def _mode_table(drug_tables: dict[str, pd.DataFrame], drug: str, mode: str) -> pd.DataFrame:
    sub = drug_tables[drug][drug_tables[drug]["fusion_mode"] == mode].copy()
    if sub.empty:
        raise ValueError(f"No rows for drug={drug}, mode={mode}")
    return sub.sort_values("rank").reset_index(drop=True)


def _rank_of(table: pd.DataFrame, keywords: list[str]) -> int | None:
    pattern = "|".join(keywords)
    matched = table[table["disease_name"].astype(str).str.contains(pattern, case=False, regex=True, na=False)]
    if matched.empty:
        return None
    return int(matched["rank"].min())


def _score_of(table: pd.DataFrame, keywords: list[str]) -> float | None:
    pattern = "|".join(keywords)
    matched = table[table["disease_name"].astype(str).str.contains(pattern, case=False, regex=True, na=False)]
    if matched.empty:
        return None
    return float(matched.sort_values("rank").iloc[0]["final_score"])


def _top3_risk(table: pd.DataFrame) -> tuple[float, float, float]:
    top3 = table.sort_values("rank").head(3)
    return (
        float(pd.to_numeric(top3["pro_inflammatory_score"], errors="coerce").fillna(0.0).mean()),
        float(pd.to_numeric(top3["toxin_risk_score"], errors="coerce").fillna(0.0).mean()),
        float(pd.to_numeric(top3["pathobiont_load"], errors="coerce").fillna(0.0).mean()),
    )


def _high_benefit_count(table: pd.DataFrame, factor: float = 0.8) -> int:
    top1 = float(table.sort_values("rank").iloc[0]["final_score"])
    if top1 <= 0:
        return 0
    return int((pd.to_numeric(table["final_score"], errors="coerce").fillna(0.0) >= factor * top1).sum())


def _profile(profile: dict[str, Any], drug: str) -> dict[str, Any]:
    if drug not in profile:
        raise KeyError(f"profile evidence missing for {drug}")
    return dict(profile[drug])


def _inhibit_fraction(p: dict[str, Any]) -> float:
    inhibit = float(dict(p.get("step1_counts", {})).get("inhibit", 0))
    panel_size = float(p.get("panel_size", 0) or 0)
    if panel_size <= 0:
        return 0.0
    return inhibit / panel_size


def _butyrate_hits(p: dict[str, Any]) -> tuple[int, int]:
    rows = p.get("butyrate_rows", []) or []
    total = len(rows)
    if total == 0:
        return 0, 0
    strong = 0
    for row in rows:
        label = str(row.get("predicted_effect_label", "")).lower()
        inhibit_prob = pd.to_numeric(pd.Series([row.get("predicted_inhibit_probability")]), errors="coerce").iloc[0]
        score = pd.to_numeric(pd.Series([row.get("predicted_effect_score")]), errors="coerce").iloc[0]
        if label == "inhibit" and ((not pd.isna(inhibit_prob) and float(inhibit_prob) >= 0.5) or (not pd.isna(score) and float(score) < 0)):
            strong += 1
    return strong, total


def _format_rank_block(table: pd.DataFrame, label: str) -> str:
    ibs = _rank_of(table, [r"IBS", r"肠易激", r"IBS-D", r"腹泻", r"Diarrhea"])
    cd = _rank_of(table, [r"CD", r"克罗恩"])
    crc = _rank_of(table, [r"CRC", r"结直肠癌"])
    ibd = _rank_of(table, [r"IBD", r"炎症性肠病"])
    abscess = _rank_of(table, [r"Abscess", r"脓肿"])
    return f"{label}: IBS={ibs}, CD={cd}, CRC={crc}, IBD={ibd}, Abscess={abscess}"


def evaluate_assertions(assertions: pd.DataFrame, drug_tables: dict[str, pd.DataFrame], profile: dict[str, Any]) -> pd.DataFrame:
    rows: list[AssertionResult] = []

    # Preload tables and profile objects.
    r_raw = _mode_table(drug_tables, "Rifaximin", RAW_MODE)
    r_mech = _mode_table(drug_tables, "Rifaximin", MECH_MODE)
    r_fin = _mode_table(drug_tables, "Rifaximin", FINAL_MODE)
    r_profile = _profile(profile, "Rifaximin")

    v_raw = _mode_table(drug_tables, "Vancomycin", RAW_MODE)
    v_mech = _mode_table(drug_tables, "Vancomycin", MECH_MODE)
    v_fin = _mode_table(drug_tables, "Vancomycin", FINAL_MODE)
    v_profile = _profile(profile, "Vancomycin")

    l_raw = _mode_table(drug_tables, "Lubiprostone", RAW_MODE)
    l_mech = _mode_table(drug_tables, "Lubiprostone", MECH_MODE)
    l_fin = _mode_table(drug_tables, "Lubiprostone", FINAL_MODE)
    l_profile = _profile(profile, "Lubiprostone")

    m_raw = _mode_table(drug_tables, "Metronidazole", RAW_MODE)
    m_mech = _mode_table(drug_tables, "Metronidazole", MECH_MODE)
    m_fin = _mode_table(drug_tables, "Metronidazole", FINAL_MODE)
    m_profile = _profile(profile, "Metronidazole")

    for _, record in assertions.iterrows():
        aid = str(record["assertion_id"])
        drug = str(record["Drug"])
        severity = str(record.get("Severity", ""))

        if aid == "RFX_01":
            raw_ibs = _rank_of(r_raw, [r"IBS", r"肠易激", r"IBS-D", r"腹泻", r"Diarrhea"])
            fin_ibs = _rank_of(r_fin, [r"IBS", r"肠易激", r"IBS-D", r"腹泻", r"Diarrhea"])
            raw_cd = _rank_of(r_raw, [r"CD", r"克罗恩"])
            fin_cd = _rank_of(r_fin, [r"CD", r"克罗恩"])
            if fin_ibs is None:
                status = "FAIL"
                support = "无"
                conflict = f"Final 未出现 IBS/IBS-D；CD rank={fin_cd}"
                confidence = "High"
            else:
                improved = raw_ibs is not None and fin_ibs < raw_ibs
                better_than_cd = fin_cd is not None and fin_ibs < fin_cd
                if improved and better_than_cd:
                    status = "PASS"
                    support = f"IBS rank improved {raw_ibs}->{fin_ibs} and better than CD({fin_cd})"
                    conflict = ""
                    confidence = "High"
                elif improved or better_than_cd:
                    status = "PARTIAL"
                    support = f"IBS rank={fin_ibs}, raw={raw_ibs}, CD={fin_cd}"
                    conflict = "未同时满足提升与炎症比较条件"
                    confidence = "Medium"
                else:
                    status = "FAIL"
                    support = ""
                    conflict = f"IBS rank={fin_ibs} not better than CD={fin_cd}"
                    confidence = "High"
            rows.append(
                AssertionResult(
                    assertion_id=aid,
                    drug=drug,
                    raw_evidence=_format_rank_block(r_raw, "raw"),
                    mechanism_evidence=_format_rank_block(r_mech, "mechanism"),
                    final_evidence=_format_rank_block(r_fin, "final"),
                    status=status,
                    confidence=confidence,
                    supporting=support,
                    conflicting=conflict,
                    notes=f"{severity}: prioritize factual IBS-D alignment over score movement.",
                )
            )
        elif aid == "RFX_02":
            strong, total = _butyrate_hits(r_profile)
            ratio = strong / total if total else 0.0
            if total == 0:
                status = "PARTIAL"
                confidence = "Low"
                support = ""
                conflict = "未匹配到核心产丁酸菌，无法完整验证"
            elif ratio <= 0.2:
                status = "PASS"
                confidence = "High"
                support = f"core butyrate strong inhibit={strong}/{total}"
                conflict = ""
            elif ratio <= 0.5:
                status = "PARTIAL"
                confidence = "Medium"
                support = f"core butyrate strong inhibit={strong}/{total}"
                conflict = "仍存在明显抑制信号"
            else:
                status = "FAIL"
                confidence = "High"
                support = ""
                conflict = f"core butyrate strong inhibit={strong}/{total} (indiscriminate pattern)"
            rows.append(
                AssertionResult(
                    assertion_id=aid,
                    drug=drug,
                    raw_evidence=f"butyrate strong_inhibit={strong}/{total}; inhibit_count={dict(r_profile.get('step1_counts', {})).get('inhibit', 0)}",
                    mechanism_evidence=f"mechanism top3 butyrate_support={r_mech.head(3)['butyrate_support_score'].mean():.4f}",
                    final_evidence=f"final top3 butyrate_support={r_fin.head(3)['butyrate_support_score'].mean():.4f}",
                    status=status,
                    confidence=confidence,
                    supporting=support,
                    conflicting=conflict,
                    notes="Guardrail focuses on core taxa-level evidence, not only disease ranks.",
                )
            )
        elif aid == "RFX_03":
            frac = _inhibit_fraction(r_profile)
            profile_tag = str(r_profile.get("drug_profile", "unknown"))
            eco_score = float(pd.to_numeric(pd.Series([r_profile.get("ecological_risk_score")]), errors="coerce").fillna(0.0).iloc[0])
            if profile_tag == "eubiotic_modulator" and frac <= 0.55 and eco_score <= 0.35:
                status = "PASS"
                confidence = "High"
                support = f"profile={profile_tag}, inhibit_fraction={frac:.2f}, eco_score={eco_score:.2f}"
                conflict = ""
            elif profile_tag == "eubiotic_modulator" and (frac <= 0.75 or eco_score <= 0.45):
                status = "PARTIAL"
                confidence = "Medium"
                support = f"profile={profile_tag}, inhibit_fraction={frac:.2f}, eco_score={eco_score:.2f}"
                conflict = "profile 已注入，但生态扰动仍偏高"
            else:
                status = "FAIL"
                confidence = "High"
                support = ""
                conflict = f"profile={profile_tag}, inhibit_fraction={frac:.2f}, eco_score={eco_score:.2f}"
            rows.append(
                AssertionResult(
                    assertion_id=aid,
                    drug=drug,
                    raw_evidence=f"step1_counts={r_profile.get('step1_counts', {})}, mean_inhibit_prob={r_profile.get('mean_predicted_inhibit_probability')}",
                    mechanism_evidence=f"mechanism risk(top3): pro/toxin/path={_top3_risk(r_mech)}",
                    final_evidence=f"final risk(top3): pro/toxin/path={_top3_risk(r_fin)}, profile={profile_tag}, eco={eco_score:.3f}",
                    status=status,
                    confidence=confidence,
                    supporting=support,
                    conflicting=conflict,
                    notes="Uses explicit drug_profile + ecological_risk evidence after gating update.",
                )
            )
        elif aid == "VAN_01":
            high_benefit = _high_benefit_count(v_fin, factor=0.8)
            risk_tuple = _top3_risk(v_fin)
            v_frac = _inhibit_fraction(v_profile)
            broad_ok = high_benefit <= 3
            risk_visible = (v_frac >= 0.30) or (risk_tuple[0] + risk_tuple[1] + risk_tuple[2] >= 0.08)
            if broad_ok and risk_visible:
                status = "PASS"
                confidence = "High"
                support = f"high_benefit_count={high_benefit}, risk_visible={risk_visible}"
                conflict = ""
            elif broad_ok or risk_visible:
                status = "PARTIAL"
                confidence = "Medium"
                support = f"high_benefit_count={high_benefit}, inhibit_fraction={v_frac:.2f}"
                conflict = f"risk channel weak in final(top3 pro+tox+path={sum(risk_tuple):.4f})"
            else:
                status = "FAIL"
                confidence = "High"
                support = ""
                conflict = f"high_benefit_count={high_benefit}, risk(top3)={risk_tuple}, inhibit_fraction={v_frac:.2f}"
            rows.append(
                AssertionResult(
                    assertion_id=aid,
                    drug=drug,
                    raw_evidence=f"raw high_benefit_count={_high_benefit_count(v_raw)}, top1={v_raw.iloc[0]['disease_name']}",
                    mechanism_evidence=f"mech high_benefit_count={_high_benefit_count(v_mech)}, risk(top3)={_top3_risk(v_mech)}",
                    final_evidence=f"final high_benefit_count={high_benefit}, risk(top3)={risk_tuple}, inhibit_fraction={v_frac:.2f}",
                    status=status,
                    confidence=confidence,
                    supporting=support,
                    conflicting=conflict,
                    notes="Evaluated with the user-corrected Vancomycin SMILES (2026-04-14).",
                )
            )
        elif aid == "VAN_02":
            cd_rank = _rank_of(v_fin, [r"CD", r"克罗恩"])
            cd_score = _score_of(v_fin, [r"CD", r"克罗恩"])
            if cd_rank is None:
                status = "PARTIAL"
                confidence = "Medium"
                support = "CD 未在当前 disease panel 出现，未见强阳性"
                conflict = "目标疾病缺失，无法做强结论"
            elif cd_rank > 3 or (cd_score is not None and cd_score < 0.15):
                status = "PASS"
                confidence = "High"
                support = f"CD rank={cd_rank}, score={cd_score:.4f}"
                conflict = ""
            else:
                status = "FAIL"
                confidence = "High"
                support = ""
                conflict = f"CD rank={cd_rank}, score={cd_score:.4f}"
            rows.append(
                AssertionResult(
                    assertion_id=aid,
                    drug=drug,
                    raw_evidence=_format_rank_block(v_raw, "raw"),
                    mechanism_evidence=_format_rank_block(v_mech, "mechanism"),
                    final_evidence=_format_rank_block(v_fin, "final"),
                    status=status,
                    confidence=confidence,
                    supporting=support,
                    conflicting=conflict,
                    notes="Disease-specific guardrail should be validated with CD-present panel in future runs.",
                )
            )
        elif aid == "LUB_01":
            frac = _inhibit_fraction(l_profile)
            const_rank = _rank_of(l_fin, [r"Constipation", r"便秘"])
            profile_tag = str(l_profile.get("drug_profile", "unknown"))
            if profile_tag == "host_secretagogue" and frac <= 0.55 and const_rank is not None and const_rank <= 4:
                status = "PASS"
                confidence = "High"
                support = f"profile={profile_tag}, inhibit_fraction={frac:.2f}, constipation_rank={const_rank}"
                conflict = ""
            elif profile_tag == "host_secretagogue" or (const_rank is not None and const_rank <= 5):
                status = "PARTIAL"
                confidence = "Medium"
                support = f"profile={profile_tag}, inhibit_fraction={frac:.2f}, constipation_rank={const_rank}"
                conflict = "host profile 已对齐，但仍有抗菌样扰动"
            else:
                status = "FAIL"
                confidence = "High"
                support = ""
                conflict = f"profile={profile_tag}, inhibit_fraction={frac:.2f}, constipation_rank={const_rank}"
            rows.append(
                AssertionResult(
                    assertion_id=aid,
                    drug=drug,
                    raw_evidence=f"raw inhibit_fraction={_inhibit_fraction(l_profile):.2f}, rank(Constipation)={_rank_of(l_raw,[r'Constipation',r'便秘'])}",
                    mechanism_evidence=f"mech barrier(top3)={l_mech.head(3)['barrier_protection_score'].mean():.4f}, pro+tox(top3)={(l_mech.head(3)['pro_inflammatory_score'].mean()+l_mech.head(3)['toxin_risk_score'].mean()):.4f}",
                    final_evidence=f"final inhibit_fraction={frac:.2f}, rank(Constipation)={const_rank}, profile={profile_tag}",
                    status=status,
                    confidence=confidence,
                    supporting=support,
                    conflicting=conflict,
                    notes="Host-pathway identity currently inferred indirectly from weak direct-microbe perturbation.",
                )
            )
        elif aid == "LUB_02":
            frac = _inhibit_fraction(l_profile)
            strong_07 = int(l_profile.get("inhibit_prob_ge_0_7", 0) or 0)
            if frac <= 0.35 and strong_07 <= 10:
                status = "PASS"
                confidence = "High"
                support = f"inhibit_fraction={frac:.2f}, inhibit>=0.7 count={strong_07}"
                conflict = ""
            elif frac <= 0.55:
                status = "PARTIAL"
                confidence = "Medium"
                support = f"inhibit_fraction={frac:.2f}, inhibit>=0.7 count={strong_07}"
                conflict = "仍有中等范围抑制"
            else:
                status = "FAIL"
                confidence = "High"
                support = ""
                conflict = f"broad antimicrobial signal: inhibit_fraction={frac:.2f}, inhibit>=0.7 count={strong_07}"
            rows.append(
                AssertionResult(
                    assertion_id=aid,
                    drug=drug,
                    raw_evidence=f"step1_counts={l_profile.get('step1_counts', {})}",
                    mechanism_evidence=f"mechanism top3 butyrate/barrier={l_mech.head(3)['butyrate_support_score'].mean():.4f}/{l_mech.head(3)['barrier_protection_score'].mean():.4f}",
                    final_evidence=f"final inhibit_fraction={frac:.2f}, inhibit>=0.7 count={strong_07}",
                    status=status,
                    confidence=confidence,
                    supporting=support,
                    conflicting=conflict,
                    notes="Guardrail checks whether explanation is dominated by antimicrobial channel.",
                )
            )
        elif aid == "MTZ_01":
            relevant_rank = min(
                value
                for value in [
                    _rank_of(m_fin, [r"Abscess", r"脓肿"]),
                    _rank_of(m_fin, [r"IBD", r"炎症性肠病"]),
                    _rank_of(m_fin, [r"CD", r"克罗恩"]),
                ]
                if value is not None
            )
            if relevant_rank <= 3:
                status = "PASS"
                confidence = "High"
                support = f"relevant_disease_best_rank={relevant_rank}"
                conflict = ""
            elif relevant_rank <= 6:
                status = "PARTIAL"
                confidence = "Medium"
                support = f"relevant_disease_best_rank={relevant_rank}"
                conflict = "benefit signal exists but not concentrated in top tier"
            else:
                status = "FAIL"
                confidence = "High"
                support = ""
                conflict = f"relevant_disease_best_rank={relevant_rank}"
            rows.append(
                AssertionResult(
                    assertion_id=aid,
                    drug=drug,
                    raw_evidence=_format_rank_block(m_raw, "raw"),
                    mechanism_evidence=_format_rank_block(m_mech, "mechanism"),
                    final_evidence=_format_rank_block(m_fin, "final"),
                    status=status,
                    confidence=confidence,
                    supporting=support,
                    conflicting=conflict,
                    notes="Contextual benefit judged on infection/inflammation subset, not global average score.",
                )
            )
        elif aid == "MTZ_02":
            frac = _inhibit_fraction(m_profile)
            risk_tuple = _top3_risk(m_fin)
            risk_visible = (frac >= 0.30) or (sum(risk_tuple) >= 0.08)
            mechanism_risk_visible = sum(_top3_risk(m_mech)) >= 0.08
            if risk_visible and mechanism_risk_visible:
                status = "PASS"
                confidence = "High"
                support = f"inhibit_fraction={frac:.2f}, mechanism risk(top3)={_top3_risk(m_mech)}"
                conflict = ""
            elif risk_visible:
                status = "PARTIAL"
                confidence = "Medium"
                support = f"raw ecological disruption visible (inhibit_fraction={frac:.2f})"
                conflict = f"mechanism risk channel weak (top3 sum={sum(_top3_risk(m_mech)):.4f})"
            else:
                status = "FAIL"
                confidence = "High"
                support = ""
                conflict = f"risk not visible: inhibit_fraction={frac:.2f}, final risk(top3)={risk_tuple}"
            rows.append(
                AssertionResult(
                    assertion_id=aid,
                    drug=drug,
                    raw_evidence=f"step1_counts={m_profile.get('step1_counts', {})}, inhibit_fraction={frac:.2f}",
                    mechanism_evidence=f"mechanism risk(top3)={_top3_risk(m_mech)}",
                    final_evidence=f"final risk(top3)={risk_tuple}",
                    status=status,
                    confidence=confidence,
                    supporting=support,
                    conflicting=conflict,
                    notes="Risk-benefit balance requires explicit risk channel calibration in mechanism layer.",
                )
            )
        else:
            rows.append(
                AssertionResult(
                    assertion_id=aid,
                    drug=drug,
                    raw_evidence="N/A",
                    mechanism_evidence="N/A",
                    final_evidence="N/A",
                    status="PARTIAL",
                    confidence="Low",
                    supporting="",
                    conflicting="No evaluator implemented for this assertion id.",
                    notes="Please extend evaluate_case_based_sanity_benchmark.py",
                )
            )

    result_df = pd.DataFrame([row.to_row() for row in rows], columns=RESULT_COLUMNS)
    return result_df


def _write_results_to_workbook(workbook_path: Path, results: pd.DataFrame) -> None:
    wb = load_workbook(workbook_path)
    ws_results = wb["Results"]

    # Clear existing rows below header.
    max_row = ws_results.max_row
    max_col = len(RESULT_COLUMNS)
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            ws_results.cell(row=row, column=col).value = None

    # Fill result rows.
    for row_idx, (_, row) in enumerate(results.iterrows(), start=2):
        for col_idx, column in enumerate(RESULT_COLUMNS, start=1):
            ws_results.cell(row=row_idx, column=col_idx).value = row[column]

    # Update Summary sheet values if present.
    ws_summary = wb["Summary"]
    status_counts = results["Status"].value_counts().to_dict()
    total = int(len(results))
    pass_count = int(status_counts.get("PASS", 0))
    partial_count = int(status_counts.get("PARTIAL", 0))
    fail_count = int(status_counts.get("FAIL", 0))
    pass_rate = pass_count / max(total, 1)
    high_sev = set(
        results.loc[
            results["assertion_id"].isin(
                {"RFX_01", "RFX_02", "VAN_01", "VAN_02", "LUB_01", "MTZ_02"}
            ),
            "assertion_id",
        ].tolist()
    )
    high_fail = int(results[(results["assertion_id"].isin(high_sev)) & (results["Status"] == "FAIL")].shape[0])
    high_partial = int(results[(results["assertion_id"].isin(high_sev)) & (results["Status"] == "PARTIAL")].shape[0])

    summary_map = {
        "Total assertions": total,
        "PASS": pass_count,
        "PARTIAL": partial_count,
        "FAIL": fail_count,
        "Pass rate": f"{pass_rate:.1%}",
        "High-severity failures": high_fail,
        "High-severity partials": high_partial,
    }
    for row in range(1, ws_summary.max_row + 1):
        key = ws_summary.cell(row=row, column=1).value
        if key in summary_map:
            ws_summary.cell(row=row, column=2).value = summary_map[key]

    wb.save(workbook_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Evaluate case-based sanity benchmark assertions and export Results.")
    parser.add_argument("--benchmark", type=Path, default=BENCHMARK_XLSX)
    parser.add_argument("--fusion-dir", type=Path, default=FUSION_DIR)
    parser.add_argument("--profile-evidence", type=Path, default=PROFILE_EVIDENCE_PATH)
    parser.add_argument("--csv-output", type=Path, default=RESULTS_CSV_PATH)
    args = parser.parse_args()

    if not args.benchmark.exists():
        raise FileNotFoundError(f"Benchmark workbook not found: {args.benchmark}")
    if not args.fusion_dir.exists():
        raise FileNotFoundError(f"Fusion directory not found: {args.fusion_dir}")
    if not args.profile_evidence.exists():
        raise FileNotFoundError(f"Profile evidence not found: {args.profile_evidence}")

    assertions = pd.read_excel(args.benchmark, sheet_name="Assertions")
    drug_tables = _load_drug_tables()
    profile = json.loads(args.profile_evidence.read_text(encoding="utf-8"))
    results = evaluate_assertions(assertions=assertions, drug_tables=drug_tables, profile=profile)

    args.csv_output.parent.mkdir(parents=True, exist_ok=True)
    results.to_csv(args.csv_output, index=False)
    _write_results_to_workbook(args.benchmark, results)

    print(
        json.dumps(
            {
                "status": "ok",
                "rows": int(len(results)),
                "csv_output": str(args.csv_output),
                "workbook": str(args.benchmark),
                "status_counts": results["Status"].value_counts().to_dict(),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
